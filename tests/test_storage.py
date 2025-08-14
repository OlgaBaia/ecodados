from pathlib import Path
from openpyxl import load_workbook
from core.storage import ensure_workbook, append_record, HEADERS
from core.models import Registro

def test_cria_e_salva(tmp_path: Path):
    path = tmp_path / "reg.xlsx"
    ensure_workbook(path)
    append_record(Registro("Ana", 30, "2025-09-01"), path)
    wb = load_workbook(path)
    ws = wb.active

    # Check header
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))[0:3]]
    assert headers == HEADERS

    # Check last row
    last = [c.value for c in ws[ws.max_row]]
    assert last == ["Ana", 30, "2025-09-01"]
