from pathlib import Path
from typing import List
from openpyxl import Workbook, load_workbook
from .models import Registro

EXCEL_PATH = Path("data") / "registros.xlsx"
HEADERS = ["nome", "idade", "data"]

def ensure_workbook(path: Path = EXCEL_PATH) -> None:
    if not path.exists():
        path.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "registros"
        ws.append(HEADERS)
        wb.save(path)

def append_record(reg: Registro, path: Path = EXCEL_PATH) -> None:
    ensure_workbook(path)
    wb = load_workbook(path)
    ws = wb.active
    ws.append([reg.nome, reg.idade, reg.data_iso])
    wb.save(path)

def list_records(path: Path = EXCEL_PATH) -> List[Registro]:
    """Lê todas as linhas (ignorando o header) e devolve como lista de Registro."""
    ensure_workbook(path)
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    out: List[Registro] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        nome, idade, data_iso = (row + (None, None, None))[:3]
        if nome is None and idade is None and data_iso is None:
            continue
        try:
            idade_int = int(idade) if idade is not None else 0
        except (TypeError, ValueError):
            continue  # pula linhas quebradas
        out.append(Registro(str(nome).strip(), idade_int, str(data_iso).strip()))
    return out
